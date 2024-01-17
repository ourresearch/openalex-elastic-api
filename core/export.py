import csv
import datetime
import io

from flask import make_response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook


def is_group_by_export(request):
    export_format = request.args.get("format")
    group_by = request.args.get("group_by") or request.args.get("group-by")
    group_bys = request.args.get("group_bys") or request.args.get("group-bys")
    return (
        export_format
        and export_format.lower() in ["csv", "xlsx"]
        and (group_by or group_bys)
    )


def export_group_by(result, request):
    group_by_key = request.args.get("group_by") or request.args.get("group-by")
    if group_by_key:
        group_by_results = format_group_by_results(result, group_by_key)
    else:
        group_by_results = format_group_bys_results(result)
    timestamp = get_timestamp()
    export_format = request.args.get("format")
    filename = f"openalex-group-by-{timestamp}.{export_format.lower()}"

    if export_format.lower() == "csv":
        response = export_group_by_csv(filename, group_by_results)
    elif export_format.lower() == "xlsx":
        response = export_group_by_xlsx(filename, group_by_results)
    else:
        raise ValueError("Invalid format")
    return response


def export_group_by_csv(filename, group_by_results):
    string_io = io.StringIO()

    max_rows = max(len(rows) for rows in group_by_results.values())

    csv_writer = csv.writer(string_io)

    top_row = []
    for group_by_key in group_by_results.keys():
        top_row.extend([friendly_header_name(group_by_key), "", ""])
    csv_writer.writerow(top_row)

    header_row = []
    for _ in group_by_results:
        header_row.extend(["name", "count", ""])
    csv_writer.writerow(header_row)

    for i in range(max_rows):
        row_data = []
        for rows in group_by_results.values():
            if i < len(rows):
                row_data.extend([rows[i]["key_display_name"], rows[i]["count"], ""])
            else:
                row_data.extend(["", "", ""])
        csv_writer.writerow(row_data)

    output = make_response(string_io.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename={filename}"
    output.headers["Content-type"] = "text/csv"
    return output


def export_group_by_xlsx(filename, group_by_results):
    wb = Workbook()
    ws = wb.active

    max_rows = max(len(rows) for rows in group_by_results.values())

    col_index = 1
    for group_by_key in group_by_results.keys():
        ws.cell(row=1, column=col_index, value=friendly_header_name(group_by_key))
        col_index += 3

    col_index = 1
    for _ in group_by_results:
        ws.cell(row=2, column=col_index, value="name")
        ws.cell(row=2, column=col_index + 1, value="count")
        col_index += 3

    for i in range(max_rows):
        col_index = 1
        for rows in group_by_results.values():
            if i < len(rows):
                ws.cell(row=i + 3, column=col_index, value=rows[i]["key_display_name"])
                ws.cell(row=i + 3, column=col_index + 1, value=rows[i]["count"])
            col_index += 3

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    output = make_response(save_virtual_workbook(wb))
    output.headers["Content-disposition"] = f"attachment; filename={filename}"
    output.headers[
        "Content-type"
    ] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return output


def format_group_by_results(result, group_by_key):
    group_by_results = {}
    group_by_results[group_by_key] = []
    for r in result["group_by"]:
        group_by_results[group_by_key].append(
            {
                "key": r["key"],
                "key_display_name": r["key_display_name"],
                "count": r["doc_count"],
            }
        )
    return group_by_results


def format_group_bys_results(result):
    group_by_results = {}

    for item in result["group_bys"]:
        group_by_key = item.get("group_by_key")
        groups = item.get("groups", [])

        if group_by_key not in group_by_results:
            group_by_results[group_by_key] = []

        for group in groups:
            group_by_results[group_by_key].append(
                {
                    "key": group["key"],
                    "key_display_name": group["key_display_name"],
                    "count": group["doc_count"],
                }
            )

    return group_by_results


def get_timestamp():
    return datetime.datetime.utcnow().strftime("%Y%m%d")


def friendly_header_name(header):
    manual_conversions = {
        "oa_status": "Open Access Status",
        "source.is_oa": "Source Is Open Access",
        "source.is_in_doaj": "Source Is In DOAJ",
        "has_pmid": "Has PubMed ID",
        "has_pmcid": "Has PubMed Central ID",
        "source.issn": "Source ISSN",
    }
    if header in manual_conversions:
        return manual_conversions[header]
    if "." in header:
        header = " ".join(header.split("."))
    return header.replace("_", " ").title()
